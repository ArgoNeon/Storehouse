import pandas as pd
import openpyxl as xl

from cell import Cell

def read_field(field_file_name):
    field_file  = xl.load_workbook(field_file_name)
    field_sheet = field_file.active

    field = []

    for irow in field_sheet.iter_rows():
        if irow[0].value is not None:
            row = []
            for icell in irow:
                if icell.value is not None:
                    row.append(icell.value)
                else:
                    row.append('b')
            field.append(row)
    
    return field

def write_field(field_file_name, field_data):
    field_file  = xl.Workbook()
    field_sheet = field_file.active

    for irow in field_data:
            row = []
            for icell in irow:
                cell_type = icell.getType()
                if (icell.isReserved()):
                    cell_type = 'res'
                if (icell.isRobot()):
                    cell_type = 'r'
                row.append(cell_type)
            field_sheet.append(row)

    field_file.save(field_file_name)
